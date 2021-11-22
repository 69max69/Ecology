from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from openpyxl.utils.cell import get_column_letter
from datetime import datetime
from .models import *

date1 = datetime.today()
date1 = date1.strftime('%d.%m.%Y.%H.%M.%S')


    # определяем стили
font_title_1 = Font(name='Times New Roman',
                size=10,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
font_title_2 = Font(name='Times New Roman',
                size=10,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

font_cap = Font(name='Times New Roman',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
font_pril = Font(name='Times New Roman',
                 size=14,
                 bold=False,
                 italic=True,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')

font_table = Font(name='Times New Roman',
                 size=14,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')


fill = PatternFill(fill_type='solid',
                       start_color='c1c1c1',
                       end_color='c2c2c2')

border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thin',
                             color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                 color='FF000000'),
                    vertical=Side(border_style='thin',
                                  color='FF000000'),
                    horizontal=Side(border_style='thin',
                                    color='FF000000')
                    )
thick_border = Border(left=Side(border_style='thick',
                              color='FF000000'),
                    right=Side(border_style='thick',
                               color='FF000000'),
                    top=Side(border_style='thick',
                             color='FF000000'),
                    bottom=Side(border_style='thick',
                                color='FF000000'),
                    diagonal=Side(border_style='thick',
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thick',
                                 color='FF000000'),
                    vertical=Side(border_style='thick',
                                  color='FF000000'),
                    horizontal=Side(border_style='thick',
                                    color='FF000000')
                    )
greyFill = PatternFill(start_color='e7e6e6',
                   end_color='e7e6e6',
                   fill_type='solid')
light_greenFill = PatternFill(start_color='92D050',
                   end_color='92D050',
                   fill_type='solid')

greenFill = PatternFill(start_color='00B050',
                   end_color='00B050',
                   fill_type='solid')

align_center = Alignment(horizontal='center',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=True,
                             shrink_to_fit=False,
                             indent=0)
align_left = Alignment(horizontal='left',
                           vertical='bottom',
                           text_rotation=0,
                           wrap_text=False,
                           shrink_to_fit=False,
                           indent=0)
number_format = 'General'
protection = Protection(locked=True,
                            hidden=False)




# Создание шапки для приложени2
def cap(sheet, year):

    # штамп
    sheet['O2'] = 'Приложение №2'
    sheet['O3'] = 'к приказу АО "Сетевая компания"'
    sheet['O4'] = 'от_________________№________"'

    #шапка
    sheet['A5'] = 'Обобщенные данные учета в области обращения с отходами '
    sheet['A6'] = 'за {} год'.format( year)
    sheet.merge_cells('A5:Q5')
    sheet.merge_cells('A6:Q6')

    # стили штампа
    for cellObj in sheet['A1:Q4']:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            sheet[cell.coordinate].font = font_pril
            sheet[cell.coordinate].alignment = align_left

    # стили шапки
    for cellObj in sheet['A5:Q6']:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            sheet[cell.coordinate].font = font_cap
            sheet[cell.coordinate].alignment = align_center

    # шапка таблицы

    sheet['A7'] = '№ строки'
    sheet.merge_cells('A7:A8')
    sheet['B7'] = 'Наименование вида отхода'
    sheet.merge_cells('B7:B8')
    sheet['C7'] = 'Код по ФККО'
    sheet.merge_cells('C7:C8')
    sheet['D7'] = 'Класс опасности вида отхода'
    sheet.merge_cells('D7:D8')
    sheet['E7'] = 'Наличие отходов на начало отчетного периода, тонн'
    sheet.merge_cells('E7:F7')
    sheet['E8'] = 'Хранение'
    sheet['F8'] = 'Накопление'

    sheet['G7'] = 'Образовано отходов в отчетном периоде, тонн'
    sheet.merge_cells('G7:G8')
    sheet['H7'] = 'Получено отходов от других лиц в отчетном периоде, тонн'
    sheet.merge_cells('H7:H8')
    sheet['I7'] = 'Обработано отходов в отчетном периоде, тонн'
    sheet.merge_cells('I7:I8')
    sheet['J7'] = 'Утилизировано отходов в отчетном периоде, тонн'
    sheet.merge_cells('J7:J8')
    sheet['K7'] = 'Обезврежено отходов в отчетном периоде, тонн'
    sheet.merge_cells('K7:K8')
    sheet['L7'] = 'Передано отходов за отчетный период, тонн'
    sheet.merge_cells('L7:L8')
    sheet['M7'] = 'Размещено отходов на эксплуатируемых объектах в отчетном периоде, тонн'
    sheet.merge_cells('M7:O7')
    sheet['M8'] = 'Всего'
    sheet['N8'] = 'Хранение'
    sheet['O8'] = 'Захоронение'

    sheet['P7'] = 'Наличие отходов на конец отчетного периода, тонн'
    sheet.merge_cells('P7:Q7')
    sheet['P8'] = 'Хранение'
    sheet['Q8'] = 'Накопление'

    sheet['A9'] = 'А'
    sheet['B9'] = '1'
    sheet['C9'] = '2'
    sheet['D9'] = '3'
    sheet['E9'] = '4'
    sheet['F9'] = '5'
    sheet['G9'] = '6'
    sheet['H9'] = '7'
    sheet['I9'] = '8'
    sheet['J9'] = '9'
    sheet['K9'] = '10'
    sheet['L9'] = '11'
    sheet['M9'] = '12'
    sheet['N9'] = '13'
    sheet['O9'] = '14'
    sheet['P9'] = '15'
    sheet['Q9'] = '16'

    sheet['B10'] = 'ВСЕГО'

    for cellObj in sheet['A7:Q9']:
        for cell in cellObj:
            sheet[cell.coordinate].font = font_table
            sheet[cell.coordinate].border = border
            sheet[cell.coordinate].alignment = align_center

    for cellObj in sheet['A10:Q10']:
        for cell in cellObj:
            sheet[cell.coordinate].font = font_cap
            sheet[cell.coordinate].border = border
            sheet[cell.coordinate].alignment = align_center
            sheet[cell.coordinate].fill = greenFill


# ширина столбцов для приложени2
def widht_sheet(sheet):
    sheet.column_dimensions[get_column_letter(1)].width = 9
    sheet.column_dimensions[get_column_letter(2)].width = 22
    sheet.column_dimensions[get_column_letter(3)].width = 17
    sheet.column_dimensions[get_column_letter(4)].width = 17
    sheet.column_dimensions[get_column_letter(5)].width = 17
    sheet.column_dimensions[get_column_letter(6)].width = 17
    sheet.column_dimensions[get_column_letter(7)].width = 17
    sheet.column_dimensions[get_column_letter(8)].width = 17
    sheet.column_dimensions[get_column_letter(9)].width = 17
    sheet.column_dimensions[get_column_letter(10)].width = 17
    sheet.column_dimensions[get_column_letter(11)].width = 17
    sheet.column_dimensions[get_column_letter(12)].width = 17
    sheet.column_dimensions[get_column_letter(13)].width = 17
    sheet.column_dimensions[get_column_letter(14)].width = 17
    sheet.column_dimensions[get_column_letter(15)].width = 17
    sheet.column_dimensions[get_column_letter(16)].width = 17
    sheet.column_dimensions[get_column_letter(17)].width = 17

# Заполнение приложения 2 по типу опасности отходов
def creat_waste_attach(sheet, danger_class, ind, educated_wastes, reclaimed_wastes, transferred_wastes, year):
    ind = ind
    wastes = Waste.objects.filter(danger_class=danger_class)
    for waste in wastes:
        educated_sum = 0
        educated_sum_before = 0
        for educated in educated_wastes:


            year_educated = educated.date_of_educated.year


            if str(educated.name) == str(waste.name):
                if  year_educated == year:
                    educated_sum += educated.quantity
                elif  year_educated <= year:
                    educated_sum_before += educated.quantity
                else:
                    pass

        reclaimed_sum = 0
        reclaimed_sum_before = 0
        for reclaimed in reclaimed_wastes:


            year_reclaimed = reclaimed.date_of_reclaimed.year


            if str(reclaimed.name) == str(waste.name):
                if  year_reclaimed == year:
                    reclaimed_sum += reclaimed.quantity
                elif  year_reclaimed <= year:
                    reclaimed_sum_before += reclaimed.quantity
                else:
                    pass

        transferred_sum = 0
        transferred_sum_before = 0
        for transferred in transferred_wastes:


            year_transferred = transferred.date_of_transferred.year


            if str(transferred.name) == str(waste.name):
                if  year_transferred == year:
                    transferred_sum += transferred.quantity
                elif  year_transferred <= year:
                    transferred_sum_before += transferred.quantity
                else:
                    pass
        accumulated = educated_sum_before - transferred_sum_before - reclaimed_sum_before
        if accumulated == 0 and educated_sum == 0 and reclaimed_sum == 0 and transferred_sum == 0:
            pass
        else:
            sheet['A' + str(10 + ind)] = waste.number
            sheet['B' + str(10 + ind)] = waste.name
            sheet['C' + str(10 + ind)] = waste.code_fkko
            sheet['D' + str(10 + ind)] = waste.danger_class

            sheet['F' + str(10 + ind)] = accumulated
            sheet['G' + str(10 + ind)] = educated_sum
            sheet['J' + str(10 + ind)] = reclaimed_sum
            sheet['L' + str(10 + ind)] = transferred_sum
            sheet['Q' + str(10 + ind)] = accumulated + educated_sum - reclaimed_sum - transferred_sum

            for cellObj in sheet['A{}:Q{}'.format(10 + ind, 10 + ind)]:
                for cell in cellObj:
                    sheet[cell.coordinate].font = font_table
                    sheet[cell.coordinate].border = border
                    sheet[cell.coordinate].alignment = align_center
            ind += 1

    return ind

#  Создание шапки для приложения 3
def cap_3(sheet, year):

    # штамп
    sheet['M2'] = 'Приложение №3'
    sheet['M3'] = 'к приказу АО "Сетевая компания"'
    sheet['M4'] = 'от_________________№________"'

    #шапка
    sheet['A5'] = 'Данные учета переданных другим лицам отходов'
    sheet['A6'] = 'за {} год'.format(year)
    sheet.merge_cells('A5:N5')
    sheet.merge_cells('A6:N6')

    # стили штампа
    for cellObj in sheet['A1:N4']:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            sheet[cell.coordinate].font = font_pril
            sheet[cell.coordinate].alignment = align_left

    # стили шапки
    for cellObj in sheet['A5:N6']:
        for cell in cellObj:
            # print(cell.coordinate, cell.value)
            sheet[cell.coordinate].font = font_cap
            sheet[cell.coordinate].alignment = align_center

    sheet['A7'] = '№ строки'
    sheet.merge_cells('A7:A8')
    sheet['B7'] = 'Наименование вида отхода'
    sheet.merge_cells('B7:B8')
    sheet['C7'] = 'Код по ФККО'
    sheet.merge_cells('C7:C8')
    sheet['D7'] = 'Класс опасности вида отхода'
    sheet.merge_cells('D7:D8')

    sheet['E7'] = 'Количество переданных отходов за отчетный период, тонн'
    sheet.merge_cells('E7:J7')
    sheet['E8'] = 'всего'
    sheet['F8'] = 'для обработки'
    sheet['G8'] = 'для утилизации'
    sheet['H8'] = 'для обезвреживания'
    sheet['I8'] = 'для хранени'
    sheet['J8'] = 'для захоронения'

    sheet['K7'] = 'Сведения о лицах, которым переданы отходы'
    sheet.merge_cells('K7:K8')
    sheet['L7'] = 'Дата и номер договора на передачу отходов'
    sheet.merge_cells('L7:L8')
    sheet['M7'] = 'Срок действия договора'
    sheet.merge_cells('M7:M8')

    sheet['N7'] = 'Реквизиты лицензии на осуществление деятельности по сбору, транспортированию, обработке, утилизации, обезвреживанию, размещению отходов I - IV классов опасности'
    sheet.merge_cells('N7:N8')


    sheet['A9'] = '1'
    sheet['B9'] = '2'
    sheet['C9'] = '3'
    sheet['D9'] = '4'
    sheet['E9'] = '5'
    sheet['F9'] = '6'
    sheet['G9'] = '7'
    sheet['H9'] = '8'
    sheet['I9'] = '9'
    sheet['J9'] = '10'
    sheet['K9'] = '11'
    sheet['L9'] = '12'
    sheet['M9'] = '13'
    sheet['N9'] = '14'



    sheet['B10'] = 'ВСЕГО'

    for cellObj in sheet['A7:N9']:
        for cell in cellObj:
            sheet[cell.coordinate].font = font_table
            sheet[cell.coordinate].border = border
            sheet[cell.coordinate].alignment = align_center

    for cellObj in sheet['A10:N10']:
        for cell in cellObj:
            sheet[cell.coordinate].font = font_cap
            sheet[cell.coordinate].border = border
            sheet[cell.coordinate].alignment = align_center
            sheet[cell.coordinate].fill = greenFill


# ширина столбцов для приложени3
def widht_sheet_3(sheet):
    sheet.column_dimensions[get_column_letter(1)].width = 9
    sheet.column_dimensions[get_column_letter(2)].width = 22
    sheet.column_dimensions[get_column_letter(3)].width = 17
    sheet.column_dimensions[get_column_letter(4)].width = 17
    sheet.column_dimensions[get_column_letter(5)].width = 17
    sheet.column_dimensions[get_column_letter(6)].width = 17
    sheet.column_dimensions[get_column_letter(7)].width = 17
    sheet.column_dimensions[get_column_letter(8)].width = 17
    sheet.column_dimensions[get_column_letter(9)].width = 17
    sheet.column_dimensions[get_column_letter(10)].width = 17
    sheet.column_dimensions[get_column_letter(11)].width = 17
    sheet.column_dimensions[get_column_letter(12)].width = 22
    sheet.column_dimensions[get_column_letter(13)].width = 17
    sheet.column_dimensions[get_column_letter(14)].width = 30

# Заполнение приложения 3 по типу опасности отходов
def creat_waste_attach_3(sheet, danger_class, ind, transferred_wastes, year):

    ind = ind
    wastes = Waste.objects.filter(danger_class=danger_class)
    for waste in wastes:
        transferred_sum = 0
        neutralization = 0 # обезвреживание
        treatment = 0 # обработка
        recycling = 0 # утилизация
        storage = 0 # хранение
        burial = 0 # захоронение
        counterparty = str
        for transferred in transferred_wastes:

            counterparty = Counterparties.objects.get(name=transferred.counterparty)


            year_transferred = transferred.date_of_transferred.year


            if str(transferred.name) == str(waste.name):
                if  year_transferred == year:
                    transferred_sum += transferred.quantity
                    if str(transferred.type_transferr) == 'для обезвреживания':
                        neutralization += transferred.quantity
                    elif str(transferred.type_transferr) == 'для обработки':
                        treatment += transferred.quantity
                    elif str(transferred.type_transferr) == 'для утилизации':
                        recycling += transferred.quantity
                    elif str(transferred.type_transferr) == 'для хранения':
                        storage += transferred.quantity
                    elif str(transferred.type_transferr) == 'для захоронения':
                        burial += transferred.quantity




        if transferred_sum == 0 and neutralization == 0 and treatment == 0 and recycling == 0 and storage == 0 and burial == 0:
            pass
        else:
            sheet['A' + str(10 + ind)] = waste.number
            sheet['B' + str(10 + ind)] = waste.name
            sheet['C' + str(10 + ind)] = waste.code_fkko
            sheet['D' + str(10 + ind)] = waste.danger_class
            sheet['E' + str(10 + ind)] = transferred_sum
            sheet['F' + str(10 + ind)] = treatment
            sheet['G' + str(10 + ind)] = recycling
            sheet['H' + str(10 + ind)] = neutralization
            sheet['I' + str(10 + ind)] = storage
            sheet['J' + str(10 + ind)] = burial
            sheet['K' + str(10 + ind)] = counterparty.name
            sheet['L' + str(10 + ind)] = '{} от {}'.format(counterparty.date_of_contract, counterparty.number_of_contract)
            sheet['M' + str(10 + ind)] = 'до {}'.format(counterparty.term_of_contract)
            sheet['N' + str(10 + ind)] = counterparty.requisites


            for cellObj in sheet['A{}:N{}'.format(10 + ind, 10 + ind)]:
                for cell in cellObj:
                    sheet[cell.coordinate].font = font_table
                    sheet[cell.coordinate].border = border
                    sheet[cell.coordinate].alignment = align_center
            ind += 1

    return ind

def creat_titel_page(obj, wb):
    sheet = wb.create_sheet()  # создание нового листа для объекта
    sheet.title = 'Титульный лист {} '.format(obj[0:10])
    sheet.column_dimensions[get_column_letter(1)].width = 1
    sheet.column_dimensions[get_column_letter(2)].width = 1
    sheet.column_dimensions[get_column_letter(3)].width = 9
    sheet.column_dimensions[get_column_letter(4)].width = 3
    sheet.column_dimensions[get_column_letter(5)].width = 6
    sheet.column_dimensions[get_column_letter(6)].width = 7
    sheet.column_dimensions[get_column_letter(7)].width = 15
    sheet.column_dimensions[get_column_letter(8)].width = 6
    sheet.column_dimensions[get_column_letter(9)].width = 23
    sheet.column_dimensions[get_column_letter(10)].width = 4
    sheet.column_dimensions[get_column_letter(11)].width = 2
    sheet.column_dimensions[get_column_letter(12)].width = 1
    sheet.column_dimensions[get_column_letter(13)].width = 15
    sheet.column_dimensions[get_column_letter(14)].width = 14
    sheet.column_dimensions[get_column_letter(15)].width = 2
    sheet.column_dimensions[get_column_letter(16)].width = 2
    sheet.column_dimensions[get_column_letter(17)].width = 1
    sheet.column_dimensions[get_column_letter(18)].width = 1
    sheet.column_dimensions[get_column_letter(19)].width = 1
    sheet.column_dimensions[get_column_letter(20)].width = 1
    sheet.column_dimensions[get_column_letter(21)].width = 1
    sheet.column_dimensions[get_column_letter(22)].width = 6
    sheet.column_dimensions[get_column_letter(23)].width = 4
    sheet.column_dimensions[get_column_letter(24)].width = 3
    sheet.column_dimensions[get_column_letter(25)].width = 3
    sheet.column_dimensions[get_column_letter(26)].width = 1
    sheet.column_dimensions[get_column_letter(27)].width = 1
    sheet.column_dimensions[get_column_letter(28)].width = 2

    sheet.row_dimensions[1].height = 14
    sheet.row_dimensions[2].height = 14
    sheet.row_dimensions[3].height = 14
    sheet.row_dimensions[4].height = 76
    sheet.row_dimensions[5].height = 14
    sheet.row_dimensions[6].height = 14
    sheet.row_dimensions[7].height = 14
    sheet.row_dimensions[8].height = 37
    sheet.row_dimensions[9].height = 14
    sheet.row_dimensions[10].height = 46
    sheet.row_dimensions[11].height = 14
    sheet.row_dimensions[12].height = 14
    sheet.row_dimensions[13].height = 14
    sheet.row_dimensions[14].height = 14
    sheet.row_dimensions[15].height = 14
    sheet.row_dimensions[16].height = 14
    sheet.row_dimensions[17].height = 14
    sheet.row_dimensions[18].height = 30
    sheet.row_dimensions[19].height = 14
    sheet.row_dimensions[20].height = 14
    sheet.row_dimensions[21].height = 14
    sheet.row_dimensions[22].height = 14
    sheet.row_dimensions[23].height = 14
    sheet.row_dimensions[24].height = 14
    sheet.row_dimensions[25].height = 14
    sheet.row_dimensions[26].height = 14
    sheet.row_dimensions[27].height = 14
    sheet.row_dimensions[28].height = 14

    sheet['F2'] = 'ДАННЫЕ УЧЕТА В ОБЛАСТИ ОБРАЩЕНИЯ С ОТХОДАМИ'
    sheet.merge_cells('F2:V2')
    sheet['F2'].fill = greyFill
    sheet['F2'].font = font_title_1
    sheet['F2'].alignment = align_center
    sheet['F2'].border = thick_border
    for cellObj in sheet['F2:V2']:
        for cell in cellObj:

            sheet[cell.coordinate].border = thick_border

    sheet['D4'] = 'Отсутствие организации учета образовавшихся, использованных, переданных другим лицам, а также размещенных отходов - нарушение п.1 ст.19 ФЗ от 24.06.1998 г. № 89-ФЗ "Об отходах производства и потребления" (влечет ответственность, установленную статьей 8.2. Кодекса Российской Федерации об административных правонарушениях). Не полный учет образовавшихся, использованных, переданных другим лицам, а также размещенных отходов - влечет ответственность, установленную статьей 8.5. Кодекса Российской Федерации об административных правонарушениях'
    sheet.merge_cells('D4:W4')
    sheet['D4'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False,indent=0)
    sheet['D4'].font = font_title_2

    for cellObj in sheet['D4:W4']:
        for cell in cellObj:

            sheet[cell.coordinate].border = thick_border




    sheet['F7'] = 'ЗАПОЛНЕНИЕ В ЭЛЕКТРОННОМ ВИДЕ'
    sheet.merge_cells('F7:V7')
    sheet['F7'].fill = greyFill
    sheet['F7'].font = font_title_2
    sheet['F7'].alignment = align_center
    for cellObj in sheet['F7:V7']:
        for cell in cellObj:
            sheet[cell.coordinate].border = thick_border

    sheet['F8'] = '(при отсутствии технической возможности ведения в электронном виде данные учета в области обращения с отходами оформляются в письменном виде)'
    sheet.merge_cells('F8:V8')

    sheet['F8'].font = font_title_2
    sheet['F8'].alignment = align_center


    sheet['G10'] = 'УЧЕТ ОБРАЗОВАВШИХСЯ, ОБРАБОТАННЫХ, УТИЛИЗИРОВАННЫХ, ОБЕЗВРЕЖЕННЫХ, ПЕРЕДАННЫХ ДРУГИМ ЛИЦАМ, А ТАКЖЕ РАЗМЕЩЕННЫХ ОТХОДОВ ПРОИЗВОДСТВА И ПОТРЕБЛЕНИЯ'
    sheet.merge_cells('G10:T10')
    sheet['G10'].font = font_title_2
    sheet['G10'].fill = greyFill
    sheet['G10'].alignment = align_center
    for cellObj in sheet['G10:T10']:
        for cell in cellObj:
            sheet[cell.coordinate].border = thick_border

    sheet['Q13'] = 'Учет отходов'
    sheet.merge_cells('Q13:AA14')
    sheet['Q13'].fill = greyFill
    sheet['Q13'].font = font_title_1
    sheet['Q13'].alignment = align_center

    for cellObj in sheet['Q13:AA14']:
        for cell in cellObj:

            sheet[cell.coordinate].border = thick_border

    sheet['Q15'] = 'Приказ МПР России:'
    sheet.merge_cells('Q15:AA15')
    sheet['Q15'].font = font_title_2
    sheet['Q15'].alignment = align_center

    sheet['Q16'] = 'Об утверждении приказа'
    sheet.merge_cells('Q16:AA16')
    sheet['Q16'].font = font_title_2
    sheet['Q16'].alignment = align_center

    sheet['Q17'] = 'от 08.12.2020 N 1028'
    sheet.merge_cells('Q17:AA17')
    sheet['Q17'].font = font_title_2
    sheet['Q17'].alignment = align_center

    sheet['Q18'] = '(действует с 01.01.2021 г. до 01.01.2027 г.)'
    sheet.merge_cells('Q18:AA18')
    sheet['Q18'].font = font_title_2
    sheet['Q18'].alignment = align_center

    sheet['P19'] = 'О внесении изменений (при наличии)'
    sheet.merge_cells('P19:AB19')
    sheet['P19'].font = font_title_2
    sheet['P19'].alignment = align_center

    for i in range(20,22):
        sheet[('S' + str(i))] = 'от'
        sheet[('S' + str(i))].font = font_title_2
        sheet.merge_cells('S{}:U{}'.format(i,i))
        sheet.merge_cells('V{}:W{}'.format(i,i))


        sheet[('X' + str(i))] = '№'
        sheet[('X' + str(i))].font = font_title_2

    sheet['Q23'] = 'Ежемесячная'
    sheet.merge_cells('Q23:AA23')
    sheet['Q23'].fill = greyFill
    sheet['Q23'].font = font_title_2
    sheet['Q23'].alignment = align_center

    for cellObj in sheet['Q23:AA23']:
        for cell in cellObj:
            sheet[cell.coordinate].border = thick_border

    sheet['A14'] = 'Предоставляют:'
    sheet.merge_cells('A14:L14')
    sheet['A14'].font = font_title_2
    sheet['A14'].alignment = align_center
    for cellObj in sheet['A14:L14']:
        for cell in cellObj:
            sheet[cell.coordinate].border = thick_border

    sheet['M14'] = 'Сроки предоставления'
    sheet.merge_cells('M14:N14')
    sheet['M14'].font = font_title_2
    sheet['M14'].alignment = align_center
    for cellObj in sheet['M14:N14']:
        for cell in cellObj:
            sheet[cell.coordinate].border = thick_border

    border_right = Border(right=Side(border_style='thin', color='FF000000'))

    sheet['B15'] = 'структурные подразделения филиала, осуществляющие деятельность в области '
    sheet.merge_cells('B15:L15')
    sheet['B15'].font = font_title_2
    for cellObj in sheet['B15:L15']:
        for cell in cellObj:
            sheet[cell.coordinate].border = border_right

    sheet['B16'] = 'обращения с отходами производства и потребления'
    sheet.merge_cells('B16:L16')
    sheet['B16'].font = font_title_2
    for cellObj in sheet['B16:L16']:
        for cell in cellObj:
            sheet[cell.coordinate].border = border_right

    for i in range(17, 24):
        if i != 19:
            sheet['L{}'.format(i)].border = border_right

    for i in range(15, 24):
        sheet['N{}'.format(i)].border = border_right

    sheet['C19'] = '- инженеру по охране окружающей среды (экологу)'
    sheet.merge_cells('C19:L19')
    sheet['C19'].font = font_title_2
    for cellObj in sheet['C19:L19']:
        for cell in cellObj:
            sheet[cell.coordinate].border = border_right

    sheet['M19'] = 'Ежемесячно'
    sheet.merge_cells('M19:N19')
    sheet['M19'].font = font_title_2
    sheet['M19'].alignment = align_center
    for cellObj in sheet['M19:N19']:
        for cell in cellObj:
            sheet[cell.coordinate].border = border_right

    sheet['M20'] = 'не позднее 25 числа месяца,'
    sheet.merge_cells('M20:N20')
    sheet['M20'].font = font_title_2
    sheet['M20'].alignment = align_center
    for cellObj in sheet['M20:N20']:
        for cell in cellObj:
            sheet[cell.coordinate].border = border_right

    sheet['M21'] = 'следующего за учетным периодом'
    sheet.merge_cells('M21:N21')
    sheet['M21'].font = font_title_2
    sheet['M21'].alignment = align_center
    for cellObj in sheet['M21:N21']:
        for cell in cellObj:
            sheet[cell.coordinate].border = border_right

    for cellObj in sheet['A23:N23']:
        for cell in cellObj:
            sheet[cell.coordinate].border = Border(bottom=Side(border_style='thin', color='FF000000'))

    sheet['L23'].border = Border(bottom=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'))
    sheet['N23'].border = Border(bottom=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'))


    for cellObj in sheet['A24:AB24']:
        for cell in cellObj:
            sheet[cell.coordinate].border = Border(bottom=Side(border_style='thin', color='FF000000'))



    for cellObj in sheet['A26:AB26']:
        for cell in cellObj:
            sheet[cell.coordinate].border = Border(bottom=Side(border_style='thin', color='FF000000'))




    for cellObj in sheet['A28:AB28']:
        for cell in cellObj:
            sheet[cell.coordinate].border = Border(bottom=Side(border_style='thin', color='FF000000'))

    sheet['B25'] = 'Наименование структурного подразделения '
    sheet.merge_cells('B25:H25')
    sheet['B25'].font = font_title_1


    sheet['I25'] = obj
    sheet['I25'].font = font_title_2
    sheet.merge_cells('I25:Y25')
    for cellObj in sheet['I25:Y25']:
        for cell in cellObj:
            sheet[cell.coordinate].border = Border(bottom=Side(border_style='thin', color='FF000000'))


    sheet['B27'] = 'Адрес'
    sheet['B27'].font = font_title_1

    objects = NameOfObjects.objects.get(name=obj)
    sheet['F27'] = objects.adress
    sheet['F27'].font = font_title_2

    for cellObj in sheet['F27:Y27']:
        for cell in cellObj:
            sheet[cell.coordinate].border = Border(bottom=Side(border_style='thin', color='FF000000'))


# Расчет итоговых сумм

def accounting_sum_2(sheet, ind_start, ind_end):

    sum_list = [0, 0, 0, 0, 0, 0, 0]
    for j in range(ind_start + 1, ind_end):
        k = 0
        for column, in 'EFGJLPQ':

            try:
                sum_list[k] = sum_list[k] + float(sheet[f'{column}{j + 10}'].value)
            except:
                pass
            k += 1
    t = 0
    for column, in 'EFGJLPQ':
        sheet[f'{column}{ind_start + 10}'] = sum_list[t]
        t += 1
    return sum_list

def accounting_sum_3(sheet, ind_start, ind_end):

    sum_list = [0, 0, 0, 0, 0, 0]
    for j in range(ind_start + 1, ind_end):
        k = 0
        for column, in 'EFGHIJ':

            try:
                sum_list[k] = sum_list[k] + float(sheet[f'{column}{j + 10}'].value)
            except:
                pass
            k += 1
    t = 0
    for column, in 'EFGHIJ':
        sheet[f'{column}{ind_start + 10}'] = sum_list[t]
        t += 1
    return sum_list





# основная функция заполнения приложений за месяц
def get_attachment_year (data):
    branch = data[0]
    user = data[1]
    year = int(data[2])

    wb = Workbook()


    # данные из БД
    objects = NameOfObjects.objects.filter(branch=branch)



    object_pk=[]
    object_list=[]
    for val in objects:
        object_pk.append(val.pk)
        object_list.append(val.name)
    ################################## сводный по филиалу ################################

    educated_wastes = EducatedWaste.objects.filter(branch=branch)

    reclaimed_wastes = ReclaimedWaste.objects.filter(branch=branch)
    transferred_wastes = TransferredWaste.objects.filter(branch=branch)

    # создание приложения 2
    sheet = wb.create_sheet()  # создание нового листа для объекта
    sheet.title = '{} Приложение 2'.format(branch)

    cap(sheet, year)  # создание шапки
    widht_sheet(sheet)
    list_all_sum = [0, 0, 0, 0, 0, 0, 0]
    ind = 1
    list_name_row = ['Всего по I классу опасности', 'Всего по II классу опасности', 'Всего по III классу опасности',
                     'Всего по IV классу опасности', 'Всего по V классу опасности']
    for i in range(0, 5):

        ind_start = ind
        sheet['A' + str(10 + ind)] = str((i + 1) * 100)
        sheet['B' + str(10 + ind)] = list_name_row[i]
        for cellObj in sheet['A{}:Q{}'.format(str(10 + ind), str(10 + ind))]:
            for cell in cellObj:
                sheet[cell.coordinate].font = font_cap
                sheet[cell.coordinate].border = border
                sheet[cell.coordinate].alignment = align_center
                sheet[cell.coordinate].fill = light_greenFill
        ind += 1

        ind_after = creat_waste_attach(sheet, (i + 1), ind, educated_wastes, reclaimed_wastes, transferred_wastes, year)

        if ind_after == ind:

            for cellObj in sheet['A{}:Q{}'.format(10 + ind, 10 + ind)]:
                for cell in cellObj:
                    sheet[cell.coordinate].font = font_table
                    sheet[cell.coordinate].border = border
                    sheet[cell.coordinate].alignment = align_center
            ind = ind_after + 1
        else:

            ind = ind_after
        ind_end = ind

        sum_list = accounting_sum_2(sheet, ind_start, ind_end)
        for i in range(len(list_all_sum)):
            list_all_sum[i] = list_all_sum[i] + sum_list[i]


    t = 0
    for column, in 'EFGJLPQ':
        sheet[f'{column}{10}'] = list_all_sum[t]
        t += 1

    # создание приложения 3
    sheet = wb.create_sheet()  # создание нового листа для объекта
    sheet.title = '{} Приложение 3'.format(branch)

    cap_3(sheet, year)  # создание шапки
    widht_sheet_3(sheet)
    list_all_sum =[0, 0, 0, 0, 0, 0]
    ind = 1
    for i in range(0, 5):

        ind_start = ind
        sheet['A' + str(10 + ind)] = str((i + 1) * 100)
        sheet['B' + str(10 + ind)] = list_name_row[i]
        for cellObj in sheet['A{}:N{}'.format(str(10 + ind), str(10 + ind))]:
            for cell in cellObj:
                sheet[cell.coordinate].font = font_cap
                sheet[cell.coordinate].border = border
                sheet[cell.coordinate].alignment = align_center
                sheet[cell.coordinate].fill = light_greenFill
        ind += 1

        ind_after = creat_waste_attach_3(sheet, (i + 1), ind, transferred_wastes, year)
        if ind_after == ind:

            for cellObj in sheet['A{}:N{}'.format(10 + ind, 10 + ind)]:
                for cell in cellObj:
                    sheet[cell.coordinate].font = font_table
                    sheet[cell.coordinate].border = border
                    sheet[cell.coordinate].alignment = align_center
            ind = ind_after + 1
        else:

            ind = ind_after
        ind_end = ind
        # заполнение сводных сумм по группе отходов
        sum_list = accounting_sum_3(sheet, ind_start, ind_end)
        for i in range(len(list_all_sum)):
            list_all_sum[i] = list_all_sum[i] + sum_list[i]
        # заполнение общей суммы
    t = 0
    for column, in 'EFGHIJ':
        sheet[f'{column}{10}'] = list_all_sum[t]
        t += 1

    ############################# по объектам ####################################
    # цикл по  объектам филиала

    for ob,pk in zip(object_list, object_pk):
        # загрузка БД по объекту
        educated_wastes = EducatedWaste.objects.filter(object_name__pk=pk)
        reclaimed_wastes = ReclaimedWaste.objects.filter(object_name__pk=pk)
        transferred_wastes = TransferredWaste.objects.filter(object_name__pk=pk)


        # создание титульного листа

        creat_titel_page(ob, wb)


        # создание приложения 2
        sheet = wb.create_sheet()# создание нового листа для объекта
        sheet.title = '{} Приложение 2'.format(ob[0:10])

        cap(sheet, year) # создание шапки
        widht_sheet(sheet)
        list_all_sum = [0, 0, 0, 0, 0, 0, 0]
        ind = 1
        list_name_row = ['Всего по I классу опасности', 'Всего по II классу опасности', 'Всего по III классу опасности',
                         'Всего по IV классу опасности', 'Всего по V классу опасности']
        for i in range(0, 5):

            ind_start = ind
            sheet['A' + str(10 + ind)] = str((i+1) * 100)
            sheet['B' + str(10 + ind)] = list_name_row[i]
            for cellObj in sheet['A{}:Q{}'.format(str(10 + ind), str(10 + ind))]:
                for cell in cellObj:
                    sheet[cell.coordinate].font = font_cap
                    sheet[cell.coordinate].border = border
                    sheet[cell.coordinate].alignment = align_center
                    sheet[cell.coordinate].fill = light_greenFill
            ind += 1

            ind_after = creat_waste_attach(sheet, (i+1), ind, educated_wastes, reclaimed_wastes, transferred_wastes, year)

            if ind_after == ind:

                for cellObj in sheet['A{}:Q{}'.format(10 + ind, 10 + ind)]:
                    for cell in cellObj:
                        sheet[cell.coordinate].font = font_table
                        sheet[cell.coordinate].border = border
                        sheet[cell.coordinate].alignment = align_center
                ind = ind_after + 1
            else:

                ind = ind_after
            ind_end = ind
# заполнение сводных сумм по группе отходов
            sum_list = accounting_sum_2(sheet, ind_start, ind_end)
            for i in range(len(list_all_sum)):
                list_all_sum[i] = list_all_sum[i] + sum_list[i]
# заполнение общей суммы
        t = 0
        for column, in 'EFGJLPQ':
            sheet[f'{column}{10}'] = list_all_sum[t]
            t += 1

        # создание приложения 3
        sheet = wb.create_sheet()  # создание нового листа для объекта
        sheet.title = '{} Приложение 3'.format(ob[0:10])

        cap_3(sheet, year)  # создание шапки
        widht_sheet_3(sheet)
        list_all_sum = [0, 0, 0, 0, 0, 0]
        ind = 1
        for i in range(0, 5):

            ind_start = ind
            sheet['A' + str(10 + ind)] = str((i+1) * 100)
            sheet['B' + str(10 + ind)] = list_name_row[i]
            for cellObj in sheet['A{}:N{}'.format(str(10 + ind), str(10 + ind))]:
                for cell in cellObj:
                    sheet[cell.coordinate].font = font_cap
                    sheet[cell.coordinate].border = border
                    sheet[cell.coordinate].alignment = align_center
                    sheet[cell.coordinate].fill = light_greenFill
            ind += 1

            ind_after = creat_waste_attach_3(sheet, (i+1), ind, transferred_wastes, year)
            if ind_after == ind:

                for cellObj in sheet['A{}:N{}'.format(10 + ind, 10 + ind)]:
                    for cell in cellObj:
                        sheet[cell.coordinate].font = font_table
                        sheet[cell.coordinate].border = border
                        sheet[cell.coordinate].alignment = align_center
                ind = ind_after + 1
            else:

                ind = ind_after
            ind_end = ind
            # заполнение сводных сумм по группе отходов
            sum_list = accounting_sum_3(sheet, ind_start, ind_end)
            for i in range(len(list_all_sum)):
                list_all_sum[i] = list_all_sum[i] + sum_list[i]
            # заполнение общей суммы
        t = 0
        for column, in 'EFGHIJ':
            sheet[f'{column}{10}'] = list_all_sum[t]
            t += 1

    filename = f'media/Act/отчет_за_{year}_год_пользователя_{user}_дата_создания_{date1}.xlsx'
    wb.save(filename)

    Report.objects.create(user=user, report=filename)